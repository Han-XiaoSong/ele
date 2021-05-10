'''
    每餐修改“发起人1号订餐人”，“time”和“hongbao”
'''
import requests
import json
import os
from openpyxl import Workbook, load_workbook
import re

nameHash = {
    "Terran": "田鹏",
    "怜子自甘心ღ": "王帅",
    "韩小松": "韩小松",
    "walter": "郑双杰",
    "chaos": "颜思行",
    "laughing": "刘华峰",
    "嘿然": "周泽林",
    "局长": "程军",
    "ZHAN": "占志虎",
    "高级灰": "郑辉",
    "佳哼": "李佳珩",
    "发起人1号订餐人": "刘华峰"
}
time = "25/11/2019"
hongbao = 2

srcUrl = 'https://h5.ele.me/spell/?cartId=cart61d9d9a78c2e46ed96606decb16331c1&sig=c442271460c68771eb60cc5ee23644f5&restaurant_id=E7149211220298205603'


def getUrlValueByKey(key="sig"):
    matchResult = re.search('(?<='+key+'=)[^&]*', srcUrl)
    return (matchResult.group())


def concatUrl():
    return 'https://h5.ele.me/restapi/booking/v1/carts/' + getUrlValueByKey('cartId')+'?sig='+getUrlValueByKey('sig')+'&from=pindan&extras[]=restaurant_info&extras[]=share_pindan&extras[]=order_status'


def getResponseJson():
    if os.path.exists("order.json"):
        with open("order.json", "r") as f:
            return json.load(f)
    else:
        print("order.json doesn't exist")
# url = 'https://h5.ele.me/restapi/booking/v1/carts/cart324d5ba94d6c41d588e456ac46ee574d?sig=ee522f6ac50c05b03af130bf0e8081c7&from=pindan&extras[]=restaurant_info&extras[]=share_pindan&extras[]=order_status&random=0.20294561891709595'
url = concatUrl()

# r = requests.get(url).json()
r = getResponseJson()

total = r["total"]
discount_amount = -r["discount_amount"] + hongbao
packing_fee = r["extra"]["packing_fee"]["price"]
agent_fee = r["extra"]["agent_fee"]["price"]
names = r["pindan"]["owners"]
order = r["group"]

if not os.path.exists("ele.xlsx"):
    wb = Workbook()
    ws = wb.create_sheet("进出表")
    ws1 = wb.create_sheet("补助表")
    wb.save("ele.xlsx")

subsidy2orderer = packing_fee + agent_fee
wb = load_workbook("ele.xlsx")
ws = wb["进出表"]
for i in range(len(names)):
    # try:
    name = nameHash[names[i]["name"]]
    # except KeyError:
        # name = nameHash["发起人1号订餐人"]
    # prename = names[i]["name"]
    # # name = nameHash[prename]
    # name = prename
    groupIndex = names[i]["group_index"]
    orders = order[groupIndex]

    originalFee = 0
    for eachOrder in orders:
        originalFee += eachOrder["total_price"]
    discount = round((originalFee * discount_amount /
                      (total - packing_fee - agent_fee)), 2)
    subsidy = 15
    if ((originalFee - discount) / 2) < 15:
        subsidy = round((originalFee - discount) / 2, 2)
    actuallyPay = round((originalFee - discount - subsidy), 2)

    row = len(ws["A"]) + 1
    ws.cell(row=row, column=2, value=time)
    ws.cell(row=row, column=3, value=actuallyPay)
    ws.cell(row=row, column=4, value=name)
    ws.cell(row=row, column=5, value=nameHash["发起人1号订餐人"])

    subsidy2orderer += subsidy
# wb.save("ele.xlsx")
# wb1 = load_workbook("ele.xlsx")
ws1 = wb["补助表"]
row1 = len(ws1["A"]) + 1
ws1.cell(row=row1, column=2, value=time)
ws1.cell(row=row1, column=3, value=subsidy2orderer)
ws1.cell(row=row1, column=4, value=nameHash["发起人1号订餐人"])
wb.save("ele.xlsx")




if __name__ == '__main__':
    print("excute Main 1")
    # getUrlValueByKey('sig')
